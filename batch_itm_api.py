#!/usr/bin/python3
#*************************************************************
# Batch job via Citrix ITM API
# Author : Jeff Hyeongjun Kim (hyeong.jun.k@gmail.com)
# *************************************************************/

import os
import sys
import json
import time
import pandas as pd
import numpy as np
from sqlalchemy import create_engine # SQL Engine to do in-memory SQL
from datetime import datetime

import citrix_itm_api as api # % request query to api must have -q 1 (quite mode to return data not print out)

## Variable
datadir = '/Users/JeffHKim/Documents/Performance/Data_Analysis'
engine = create_engine('sqlite://', echo=False)  # Create an in-memory SQLite database.
dateTimeNow = datetime.now()

## Function
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
    filename : File path or existing ExcelWriter
                (Example: '/path/to/file.xlsx')
    df : dataframe to save to workbook
    sheet_name : Name of sheet which will contain DataFrame.
                    (default: 'Sheet1')
    startrow : upper left cell row to dump data frame.
                Per default (startrow=None) calculate the last row
                in the existing DF and write to the next row...
    truncate_sheet : truncate (remove and recreate) [sheet_name]
                        before writing DataFrame to Excel file
    to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

## Job 1: Performance data for Cloud Platform on last 1 week
"""
    Target Countries: Singapore(SG), Malaysia(MY), Indonesia(ID), Thailand(TH), Vietnam(VN), Cambodia(KH), Taiwan(TW), Philippines(PH)
         0    1   2            3     4
0    4  217  TW       Taiwan  3.58
1   17  231  VN     Viet Nam  1.06
2   35  169  PH  Philippines  0.38
3   36  188  SG    Singapore  0.37
4   42  207  TH     Thailand  0.30
5   45  149  MY     Malaysia  0.23
6  116  110  KH     Cambodia  0.01
"""
#target_term = 'last_7_days'
#target_term = 'last_1_days'
target_term = '2020-04-20T00:00:00Z/2020-04-26T23:59:59Z'

""" Competitors """
competitors_df = api.main(['citrix_itm_api.py', '-d', 'platforms', '-qr', 'SELECT * FROM data_df WHERE category = "Cloud Computing" AND is_public = 1 ORDER BY id ASC', '-q', '1'])
#competitors_df = api.main(['citrix_itm_api.py', '-d', 'platforms', '-qr', 'SELECT * FROM data_df WHERE id = 43184 OR id = 43185 OR id = 43186 ', '-q', '1'])

""" Country """
#country_df = api.main(['citrix_itm_api.py', '-d', 'countries', '-qr', 'SELECT * FROM data_df WHERE iso in ("SG","MY","ID","TH","VN","KH","TW","PH")', '-q', '1'])
country_df = api.main(['citrix_itm_api.py', '-d', 'countries', '-qr', 'SELECT * FROM data_df', '-q', '1'])
http_response = []
availability = []
throughput = []
for competitor in competitors_df.itertuples():
    no, pid, name = competitor[0], competitor[2], competitor[3]

    """ 
    Run API to get performance data Median, Availability on Last 1 week
    HTTP Response Time: citrix_itm_api.py -d radar -p platformId -tr last_7_days -r "countryId,percentile_25,percentile_50,percentile_75,percentile_95,mean"
    Availability: citrix_itm_api.py -d radar -p platformId -tr last_7_days -r "countryId,availability"

    result => type: <class 'dict'>
    """
    api_query = [
        'citrix_itm_api.py', '-d', 'radar', '-p', str(pid), '-tr', target_term, '-r', 
        'countryId,percentile_10,percentile_25,percentile_50,percentile_75,percentile_90,percentile_95,mean,std_dev', '-q', '1'
    ]
    tmp_http_response = api.main(api_query)
    for item in tmp_http_response['facts']:
        http_response.append([pid, name, item[0], item[1], item[2], item[3], item[4], item[5], item[6], item[7], item[8]])

    api_query = [
        'citrix_itm_api.py', '-d', 'radar', '-p', str(pid), '-tr', target_term, '-t', '1', # ProbeType: 1 e.g. -t 1
        '-r', 'countryId,availability', '-q', '1'
    ]
    tmp_availability = api.main(api_query)
    for item in tmp_availability['facts']:
        availability.append([pid, name, item[0], item[1]])

    """api_query = [
        'citrix_itm_api.py', '-d', 'radar', '-p', str(pid), '-tr', target_term, '-t', '14', # ProbeType: 1 e.g. -t 1
        '-r', 'countryId,percentile_10,percentile_25,percentile_50,percentile_75,percentile_90,percentile_95,mean,std_dev', '-q', '1'
    ]        
    tmp_throughput = api.main(api_query)
    for item in tmp_throughput['facts']:
        throughput.append([pid, name, item[0], item[1], item[2], item[3], item[4], item[5]])"""

    print('[{}] Completd {}:{} - {}'.format(no, pid, name, dateTimeNow.strftime('%m%d_%H%M')))
    time.sleep(300) # Add delay 300 sec

# Data Processing via SQLite
competitors_df.columns = ['no', 'pid', 'competitor', 'category', 'is_public']
competitors_df = competitors_df.to_sql('competitors_df',engine) # Create SQL Table for each data : % Data should be type Dataframe

country_df.columns = ['no', 'cid', 'iso', 'country', 'ratio']
country_df = country_df.to_sql('country_df',engine)

http_response_df = pd.DataFrame(http_response, columns=['pid', 'name', 'cid', 'percentile_10', 'percentile_25', 'percentile_50', 'percentile_75', 'percentile_90', 'percentile_95', 'mean', 'std_dev'])
http_response_df = http_response_df.to_sql('http_response_df',engine)

availability_df = pd.DataFrame(availability, columns=['pid', 'name', 'cid', 'availability'])
availability_df = availability_df.to_sql('availability_df',engine)

"""throughput_df = pd.DataFrame(throughput, columns=['pid', 'name', 'cid', 'percentile_10', 'percentile_25', 'percentile_50', 'percentile_75', 'percentile_90', 'percentile_95', 'mean', 'std_dev'])
throughput_df = throughput_df.to_sql('throughput_df',engine)"""

query = """
SELECT 
    CN.iso AS country_code, CN.country AS country_name, C.competitor AS competitor, percentile_10, percentile_25, percentile_50, percentile_75, percentile_90, percentile_95, mean, std_dev, availability
FROM http_response_df AS H
LEFT OUTER JOIN availability_df AS A ON H.cid = A.cid AND H.pid = A.pid
LEFT OUTER JOIN competitors_df AS C ON C.pid = H.pid
LEFT OUTER JOIN country_df AS CN ON H.cid = CN.cid
ORDER BY CN.iso ASC"""
result = engine.execute(query).fetchall()  # List Type
data_df = pd.DataFrame(result, columns=['country_code', 'country_name', 'competitor', 'p10', 'p25', 'median', 'p75', 'p90', 'p95', 'mean', 'std_dev', 'availability'])

# Store data to Excel File
append_df_to_excel('{}/ecp_performance_itm.xlsx'.format(datadir), data_df, sheet_name=dateTimeNow.strftime('%m%d_%H%M'), index=False)

print('Done')